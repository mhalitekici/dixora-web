from __future__ import annotations

import csv
import io
import re
import unicodedata
import zipfile
from collections.abc import Iterable
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]
from openpyxl.formatting.rule import FormulaRule  # type: ignore[import-untyped]
from openpyxl.styles import (  # type: ignore[import-untyped]
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
from openpyxl.worksheet.datavalidation import DataValidation  # type: ignore[import-untyped]
from openpyxl.worksheet.table import Table, TableStyleInfo  # type: ignore[import-untyped]

from app.schemas import ProductCsvImportError

MAX_PRODUCT_CSV_BYTES = 2 * 1024 * 1024
MAX_PRODUCT_XLSX_BYTES = 5 * 1024 * 1024
MAX_PRODUCT_CSV_ROWS = 2_000
MAX_XLSX_UNCOMPRESSED_BYTES = 25 * 1024 * 1024
MAX_XLSX_ARCHIVE_MEMBERS = 200
EXCEL_SEPARATOR_DIRECTIVE = "sep=;"

TEMPLATE_HEADERS = (
    "kategori",
    "urun_adi",
    "satis_fiyati",
    "dahili_ad",
    "aciklama",
    "sku",
    "maliyet_fiyati",
    "kdv_orani",
    "hazirlama_dakika",
    "istasyon",
    "aktif",
    "satisa_acik",
    "qr_menude_goster",
    "garson_menusunde_goster",
    "stok_takibi",
)

XLSX_TEMPLATE_HEADERS = (
    "Kategori",
    "Ürün Adı",
    "Satış Fiyatı",
    "Dahili Ad",
    "Açıklama",
    "SKU",
    "Maliyet Fiyatı",
    "KDV Oranı",
    "Hazırlama Dakika",
    "İstasyon",
    "Aktif",
    "Satışa Açık",
    "QR Menüde Göster",
    "Garson Menüsünde Göster",
    "Stok Takibi",
)

_SAMPLE_ROW: tuple[object, ...] = (
    "Ana Yemekler",
    "Izgara Köfte",
    325,
    "",
    "Pilav ve mevsim salatası ile",
    "YEM-001",
    120,
    10,
    15,
    "Mutfak",
    "evet",
    "evet",
    "evet",
    "evet",
    "hayır",
)

_ALIASES = {
    "kategori": "category",
    "category": "category",
    "kategori_adi": "category",
    "urun_adi": "name",
    "urun": "name",
    "name": "name",
    "product_name": "name",
    "satis_fiyati": "selling_price",
    "fiyat": "selling_price",
    "selling_price": "selling_price",
    "price": "selling_price",
    "dahili_ad": "internal_name",
    "internal_name": "internal_name",
    "aciklama": "description",
    "description": "description",
    "sku": "sku",
    "maliyet_fiyati": "cost_price",
    "maliyet": "cost_price",
    "cost_price": "cost_price",
    "kdv_orani": "tax_rate",
    "kdv": "tax_rate",
    "tax_rate": "tax_rate",
    "hazirlama_dakika": "preparation_minutes",
    "hazirlama_suresi": "preparation_minutes",
    "preparation_minutes": "preparation_minutes",
    "istasyon": "station",
    "station": "station",
    "aktif": "is_active",
    "is_active": "is_active",
    "satisa_acik": "is_available",
    "is_available": "is_available",
    "qr_menude_goster": "qr_visible",
    "qr_visible": "qr_visible",
    "garson_menusunde_goster": "waiter_visible",
    "waiter_visible": "waiter_visible",
    "stok_takibi": "track_inventory",
    "track_inventory": "track_inventory",
}

_TRUE_VALUES = {"1", "true", "evet", "e", "yes", "y", "aktif"}
_FALSE_VALUES = {"0", "false", "hayir", "h", "no", "n", "pasif"}


@dataclass(slots=True)
class ParsedProductCsvRow:
    row_number: int
    category: str
    name: str
    selling_price: Decimal
    internal_name: str | None = None
    description: str | None = None
    sku: str | None = None
    cost_price: Decimal = Decimal("0.00")
    tax_rate: Decimal = Decimal("0.00")
    preparation_minutes: int | None = None
    station: str | None = None
    is_active: bool = True
    is_available: bool = True
    qr_visible: bool = True
    waiter_visible: bool = True
    track_inventory: bool = False


@dataclass(slots=True)
class ParsedProductCsv:
    total_rows: int = 0
    rows: list[ParsedProductCsvRow] = field(default_factory=list)
    errors: list[ProductCsvImportError] = field(default_factory=list)


def build_product_csv_template() -> bytes:
    output = io.StringIO(newline="")
    # Excel uses this directive to open the CSV in separate columns regardless
    # of the workstation's regional list-separator setting. UTF-8 BOM keeps
    # Turkish characters intact in Windows Excel.
    output.write(f"{EXCEL_SEPARATOR_DIRECTIVE}\r\n")
    writer = csv.writer(output, delimiter=";", lineterminator="\r\n")
    writer.writerow(TEMPLATE_HEADERS)
    writer.writerow(
        (
            "Ana Yemekler",
            "Izgara Köfte",
            "325,00",
            "",
            "Pilav ve mevsim salatası ile",
            "YEM-001",
            "120,00",
            "10",
            "15",
            "Mutfak",
            "evet",
            "evet",
            "evet",
            "evet",
            "hayır",
        )
    )
    return ("\ufeff" + output.getvalue()).encode("utf-8")


def build_product_xlsx_template() -> bytes:
    """Build a native Excel workbook with a fillable product table.

    XLSX stores strings as Unicode, so Turkish characters survive without any
    workstation-specific CSV encoding or delimiter settings.
    """

    workbook = Workbook()
    products = workbook.active
    products.title = "Ürünler"
    products.sheet_view.showGridLines = False
    products.freeze_panes = "A2"

    products.append(XLSX_TEMPLATE_HEADERS)
    products.append(_SAMPLE_ROW)
    table = Table(displayName="DixoraUrunTablosu", ref=f"A1:O{MAX_PRODUCT_CSV_ROWS + 1}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    products.add_table(table)

    header_fill = PatternFill("solid", fgColor="172033")
    required_fill = PatternFill("solid", fgColor="E45124")
    header_font = Font(color="FFFFFF", bold=True)
    for column_index, cell in enumerate(products[1], start=1):
        cell.fill = required_fill if column_index <= 3 else header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
    products.row_dimensions[1].height = 28
    products.row_dimensions[2].height = 30

    widths = (22, 28, 16, 22, 42, 16, 17, 13, 19, 18, 12, 16, 19, 24, 15)
    for column_index, width in enumerate(widths, start=1):
        products.column_dimensions[get_column_letter(column_index)].width = width
    products["C2"].number_format = "#,##0.00 [$₺-tr-TR]"
    products["G2"].number_format = "#,##0.00 [$₺-tr-TR]"
    products["H2"].number_format = "0.00"
    products["E2"].alignment = Alignment(wrap_text=True, vertical="top")

    yes_no_validation = DataValidation(
        type="list",
        formula1='"evet,hayır"',
        allow_blank=True,
        error="Yalnızca evet veya hayır seçin.",
        errorTitle="Geçersiz değer",
    )
    yes_no_validation.prompt = "Açılır listeden evet veya hayır seçin."
    yes_no_validation.promptTitle = "Evet / hayır"
    yes_no_validation.showErrorMessage = True
    yes_no_validation.showInputMessage = True
    products.add_data_validation(yes_no_validation)
    yes_no_validation.add(f"K2:O{MAX_PRODUCT_CSV_ROWS + 1}")

    non_negative_decimal = DataValidation(
        type="decimal",
        operator="greaterThanOrEqual",
        formula1="0",
        allow_blank=True,
        error="Sıfır veya daha büyük bir sayı girin.",
        errorTitle="Geçersiz sayı",
    )
    non_negative_decimal.showErrorMessage = True
    products.add_data_validation(non_negative_decimal)
    for column in ("C", "G", "H"):
        non_negative_decimal.add(f"{column}2:{column}{MAX_PRODUCT_CSV_ROWS + 1}")

    minutes_validation = DataValidation(
        type="whole",
        operator="between",
        formula1="0",
        formula2="1440",
        allow_blank=True,
        error="0 ile 1440 arasında tam sayı girin.",
        errorTitle="Geçersiz süre",
    )
    minutes_validation.showErrorMessage = True
    products.add_data_validation(minutes_validation)
    minutes_validation.add(f"I2:I{MAX_PRODUCT_CSV_ROWS + 1}")

    # Warn visually when one of the three required cells is empty. Formula
    # rules do not add thousands of styled cells to the workbook.
    missing_fill = PatternFill("solid", fgColor="FDE7DF")
    for column in ("A", "B", "C"):
        products.conditional_formatting.add(
            f"{column}2:{column}{MAX_PRODUCT_CSV_ROWS + 1}",
            FormulaRule(formula=[f'LEN(TRIM({column}2&""))=0'], fill=missing_fill),
        )

    guide = workbook.create_sheet("Nasıl Kullanılır")
    _build_xlsx_guide(guide)
    workbook.active = 0
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def parse_product_csv(data: bytes) -> ParsedProductCsv:
    parsed = ParsedProductCsv()
    if not data:
        parsed.errors.append(_error(1, None, "CSV dosyası boş."))
        return parsed
    if len(data) > MAX_PRODUCT_CSV_BYTES:
        parsed.errors.append(
            _error(
                1,
                None,
                f"CSV dosyası en fazla {MAX_PRODUCT_CSV_BYTES // 1024 // 1024} MB olabilir.",
            )
        )
        return parsed
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        parsed.errors.append(_error(1, None, "CSV dosyası UTF-8 kodlamasında olmalıdır."))
        return parsed
    if "\x00" in text:
        parsed.errors.append(_error(1, None, "CSV dosyası geçersiz karakter içeriyor."))
        return parsed

    csv_text, declared_delimiter, line_offset = _extract_excel_separator(text)
    if declared_delimiter is not None:
        delimiter = declared_delimiter
    else:
        try:
            delimiter = csv.Sniffer().sniff(csv_text[:8192], delimiters=";,\t").delimiter
        except csv.Error:
            delimiter = ";"

    reader = csv.reader(io.StringIO(csv_text, newline=""), delimiter=delimiter)
    try:
        raw_headers = next(reader)
    except StopIteration:
        parsed.errors.append(_error(1 + line_offset, None, "CSV dosyası başlık satırı içermiyor."))
        return parsed

    headers: list[str | None] = []
    seen: set[str] = set()
    for raw_header in raw_headers:
        canonical = _ALIASES.get(_normalize(raw_header))
        headers.append(canonical)
        if canonical is not None:
            if canonical in seen:
                parsed.errors.append(
                    _error(
                        1 + line_offset,
                        canonical,
                        f"'{raw_header.strip()}' sütunu birden fazla kez tanımlanmış.",
                    )
                )
            seen.add(canonical)
    missing = {"category", "name", "selling_price"} - seen
    for field_name in sorted(missing):
        parsed.errors.append(
            _error(
                1 + line_offset,
                field_name,
                f"Zorunlu sütun eksik: {_display_name(field_name)}.",
            )
        )
    if parsed.errors:
        return parsed

    non_empty_rows = 0
    for row_number, raw_values in enumerate(reader, start=2 + line_offset):
        if not any(value.strip() for value in raw_values):
            continue
        non_empty_rows += 1
        parsed.total_rows = non_empty_rows
        if non_empty_rows > MAX_PRODUCT_CSV_ROWS:
            parsed.errors.append(
                _error(
                    row_number,
                    None,
                    f"Tek seferde en fazla {MAX_PRODUCT_CSV_ROWS} ürün aktarılabilir.",
                )
            )
            break
        values = {
            header: raw_values[index].strip()
            for index, header in enumerate(headers)
            if header is not None and index < len(raw_values)
        }
        row, errors = _parse_row(row_number, values, delimiter)
        parsed.errors.extend(errors)
        if row is not None:
            parsed.rows.append(row)
    if non_empty_rows == 0:
        parsed.errors.append(
            _error(2 + line_offset, None, "CSV dosyasında aktarılacak ürün bulunamadı.")
        )
    return parsed


def parse_product_xlsx(data: bytes) -> ParsedProductCsv:
    parsed = ParsedProductCsv()
    if not data:
        parsed.errors.append(_error(1, None, "Excel dosyası boş."))
        return parsed
    if len(data) > MAX_PRODUCT_XLSX_BYTES:
        parsed.errors.append(
            _error(
                1,
                None,
                f"Excel dosyası en fazla {MAX_PRODUCT_XLSX_BYTES // 1024 // 1024} MB olabilir.",
            )
        )
        return parsed
    archive_error = _validate_xlsx_archive(data)
    if archive_error is not None:
        parsed.errors.append(_error(1, None, archive_error))
        return parsed

    try:
        workbook = load_workbook(
            filename=io.BytesIO(data),
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except (OSError, ValueError, KeyError, zipfile.BadZipFile):
        parsed.errors.append(
            _error(1, None, "Excel dosyası okunamadı. Geçerli bir .xlsx dosyası yükleyin.")
        )
        return parsed

    try:
        sheet_and_header = _find_xlsx_product_sheet(workbook.worksheets)
        if sheet_and_header is None:
            parsed.errors.append(
                _error(
                    1,
                    None,
                    "Excel dosyasında Kategori, Ürün Adı ve Satış Fiyatı başlıklarını "
                    "içeren bir tablo bulunamadı.",
                )
            )
            return parsed
        worksheet, header_row_number, raw_headers = sheet_and_header
        headers, header_errors = _map_headers(raw_headers, header_row_number)
        parsed.errors.extend(header_errors)
        if header_errors:
            return parsed

        non_empty_rows = 0
        last_allowed_row = header_row_number + MAX_PRODUCT_CSV_ROWS
        if worksheet.max_row > last_allowed_row:
            parsed.errors.append(
                _error(
                    last_allowed_row + 1,
                    None,
                    f"Excel tablosu en fazla {MAX_PRODUCT_CSV_ROWS} ürün satırı içerebilir.",
                )
            )
        for row_number, cells in enumerate(
            worksheet.iter_rows(
                min_row=header_row_number + 1,
                max_row=min(worksheet.max_row, last_allowed_row),
                max_col=len(headers),
            ),
            start=header_row_number + 1,
        ):
            raw_values: list[str] = []
            formula_errors: list[ProductCsvImportError] = []
            for column_index, cell in enumerate(cells[: len(headers)]):
                if cell.data_type == "f":
                    field_name = headers[column_index]
                    formula_errors.append(
                        _error(
                            row_number,
                            field_name,
                            "Formül kullanılamaz; hesaplanan sonucu sabit değer olarak girin.",
                        )
                    )
                    raw_values.append("")
                else:
                    raw_values.append(_xlsx_value_to_text(cell.value))
            if not any(value.strip() for value in raw_values) and not formula_errors:
                continue
            non_empty_rows += 1
            parsed.total_rows = non_empty_rows
            if non_empty_rows > MAX_PRODUCT_CSV_ROWS:
                parsed.errors.append(
                    _error(
                        row_number,
                        None,
                        f"Tek seferde en fazla {MAX_PRODUCT_CSV_ROWS} ürün aktarılabilir.",
                    )
                )
                break
            values = {
                header: raw_values[index].strip()
                for index, header in enumerate(headers)
                if header is not None and index < len(raw_values)
            }
            row, row_errors = _parse_row(row_number, values, ";")
            parsed.errors.extend(formula_errors)
            parsed.errors.extend(row_errors)
            if row is not None and not formula_errors:
                parsed.rows.append(row)
        if non_empty_rows == 0:
            parsed.errors.append(
                _error(
                    header_row_number + 1,
                    None,
                    "Excel tablosunda aktarılacak ürün bulunamadı.",
                )
            )
        return parsed
    finally:
        workbook.close()


def parse_product_import(data: bytes, filename: str) -> ParsedProductCsv:
    if filename.casefold().endswith(".xlsx"):
        return parse_product_xlsx(data)
    return parse_product_csv(data)


def _validate_xlsx_archive(data: bytes) -> str | None:
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            members = archive.infolist()
            if len(members) > MAX_XLSX_ARCHIVE_MEMBERS:
                return "Excel dosyası beklenenden fazla bölüm içeriyor."
            if any(member.flag_bits & 0x1 for member in members):
                return "Parola korumalı Excel dosyaları desteklenmiyor."
            if sum(member.file_size for member in members) > MAX_XLSX_UNCOMPRESSED_BYTES:
                return "Excel dosyasının açılmış boyutu izin verilen sınırı aşıyor."
            workbook_names = {member.filename.casefold() for member in members}
            if "xl/workbook.xml" not in workbook_names:
                return "Geçerli bir .xlsx çalışma kitabı yükleyin."
    except zipfile.BadZipFile:
        return "Excel dosyası okunamadı. Geçerli bir .xlsx dosyası yükleyin."
    return None


def _find_xlsx_product_sheet(
    worksheets: Iterable[Any],
) -> tuple[Any, int, list[str]] | None:
    ordered_worksheets = sorted(
        worksheets,
        key=lambda sheet: (sheet.title.casefold() not in {"ürünler", "urunler"}, sheet.title),
    )
    for worksheet in ordered_worksheets:
        for row_number, cells in enumerate(
            worksheet.iter_rows(
                min_row=1,
                max_row=min(worksheet.max_row, 20),
                max_col=min(worksheet.max_column, 100),
            ),
            start=1,
        ):
            raw_headers = [_xlsx_value_to_text(cell.value) for cell in cells]
            canonical = {_ALIASES.get(_normalize(header)) for header in raw_headers}
            if {"category", "name", "selling_price"}.issubset(canonical):
                while raw_headers and not raw_headers[-1].strip():
                    raw_headers.pop()
                return worksheet, row_number, raw_headers
    return None


def _xlsx_value_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "evet" if value else "hayır"
    if isinstance(value, float):
        return format(Decimal(str(value)), "f")
    return str(value)


def _map_headers(
    raw_headers: Iterable[str],
    header_row_number: int,
) -> tuple[list[str | None], list[ProductCsvImportError]]:
    headers: list[str | None] = []
    errors: list[ProductCsvImportError] = []
    seen: set[str] = set()
    for raw_header in raw_headers:
        canonical = _ALIASES.get(_normalize(raw_header))
        headers.append(canonical)
        if canonical is not None:
            if canonical in seen:
                errors.append(
                    _error(
                        header_row_number,
                        canonical,
                        f"'{raw_header.strip()}' sütunu birden fazla kez tanımlanmış.",
                    )
                )
            seen.add(canonical)
    missing = {"category", "name", "selling_price"} - seen
    for field_name in sorted(missing):
        errors.append(
            _error(
                header_row_number,
                field_name,
                f"Zorunlu sütun eksik: {_display_name(field_name)}.",
            )
        )
    return headers, errors


def _build_xlsx_guide(guide: Any) -> None:
    guide.sheet_view.showGridLines = False
    guide.merge_cells("A1:E1")
    guide["A1"] = "Dixora ürün aktarım şablonu"
    guide["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    guide["A1"].fill = PatternFill("solid", fgColor="172033")
    guide["A1"].alignment = Alignment(vertical="center")
    guide.row_dimensions[1].height = 38

    guide.merge_cells("A3:E3")
    guide["A3"] = "1. Ürünler sekmesine geçin ve örnek satırı kendi ürününüzle değiştirin."
    guide.merge_cells("A4:E4")
    guide["A4"] = "2. Her ürün için tek satır kullanın; başlıkları silmeyin veya değiştirmeyin."
    guide.merge_cells("A5:E5")
    guide["A5"] = "3. Dosyayı .xlsx biçiminde kaydedip Dixora'da önizleyerek içe aktarın."
    guide.merge_cells("A6:E6")
    guide["A6"] = (
        "Turuncu başlıklar zorunludur. Kategori adı paneldeki aktif kategoriyle aynı olmalıdır."
    )
    for row_number in range(3, 7):
        guide.cell(row=row_number, column=1).alignment = Alignment(wrap_text=True)
    guide["A6"].font = Font(bold=True, color="B93815")

    headers = ("Sütun", "Zorunlu", "Ne girilmeli?", "Örnek", "Not")
    guide.append(())
    guide.append(headers)
    field_rows = (
        ("Kategori", "Evet", "Paneldeki aktif kategori adı", "Ana Yemekler", "Birebir eşleşir"),
        (
            "Ürün Adı",
            "Evet",
            "Müşterinin göreceği ürün adı",
            "Izgara Köfte",
            "En fazla 160 karakter",
        ),
        ("Satış Fiyatı", "Evet", "Vergi dahil satış fiyatı", "325,00", "Negatif olamaz"),
        ("Dahili Ad", "Hayır", "Ekip içinde kullanılan kısa ad", "Köfte porsiyon", ""),
        ("Açıklama", "Hayır", "QR menü ürün açıklaması", "Pilav ve salata ile", ""),
        ("SKU", "Hayır", "İşletmede benzersiz stok kodu", "YEM-001", "Tekrar edemez"),
        ("Maliyet Fiyatı", "Hayır", "Ürünün işletmeye maliyeti", "120,00", "Boşsa 0"),
        ("KDV Oranı", "Hayır", "0–100 arası KDV yüzdesi", "10", "Boşsa 0"),
        ("Hazırlama Dakika", "Hayır", "Tahmini hazırlama süresi", "15", "0–1440 tam sayı"),
        ("İstasyon", "Hayır", "Aktif hazırlama istasyonu adı", "Mutfak", "Birebir eşleşir"),
        ("Aktif", "Hayır", "Ürün kayıt durumu", "evet", "evet / hayır"),
        ("Satışa Açık", "Hayır", "Siparişe eklenebilir mi?", "evet", "evet / hayır"),
        ("QR Menüde Göster", "Hayır", "QR menüde yayınlansın mı?", "evet", "evet / hayır"),
        (
            "Garson Menüsünde Göster",
            "Hayır",
            "Garson ekranında görünsün mü?",
            "evet",
            "evet / hayır",
        ),
        ("Stok Takibi", "Hayır", "Stok hareketi oluşturulsun mu?", "hayır", "evet / hayır"),
    )
    for row in field_rows:
        guide.append(row)
    guide_table = Table(displayName="DixoraAlanAciklamalari", ref="A8:E23")
    guide_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    guide.add_table(guide_table)
    guide.freeze_panes = "A9"
    for column, width in zip(("A", "B", "C", "D", "E"), (27, 13, 42, 24, 25), strict=True):
        guide.column_dimensions[column].width = width
    thin = Side(style="thin", color="D9DEE8")
    for row in guide.iter_rows(min_row=8, max_row=23, min_col=1, max_col=5):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(bottom=thin)


def _extract_excel_separator(text: str) -> tuple[str, str | None, int]:
    """Remove an optional Excel ``sep=<char>`` directive.

    Only Excel's comma and semicolon directives are accepted. Other
    first lines remain ordinary CSV content and therefore fail normal header
    validation instead of influencing parsing.
    """

    first_line, separator, remainder = text.partition("\n")
    normalized = first_line.removesuffix("\r").strip().casefold()
    if len(normalized) == 5 and normalized.startswith("sep="):
        delimiter = normalized[-1]
        if delimiter in {";", ","}:
            return (remainder if separator else "", delimiter, 1)
    return text, None, 0


def _parse_row(
    row_number: int,
    values: dict[str, str],
    delimiter: str,
) -> tuple[ParsedProductCsvRow | None, list[ProductCsvImportError]]:
    errors: list[ProductCsvImportError] = []
    category = values.get("category", "").strip()
    name = values.get("name", "").strip()
    if not category:
        errors.append(_error(row_number, "category", "Kategori boş bırakılamaz."))
    if not name:
        errors.append(_error(row_number, "name", "Ürün adı boş bırakılamaz."))
    elif len(name) > 160:
        errors.append(_error(row_number, "name", "Ürün adı en fazla 160 karakter olabilir."))
    if len(category) > 120:
        errors.append(
            _error(row_number, "category", "Kategori adı en fazla 120 karakter olabilir.")
        )

    selling_price = _decimal_value(
        values.get("selling_price", ""), row_number, "selling_price", delimiter, errors
    )
    cost_price = _decimal_value(
        values.get("cost_price") or "0", row_number, "cost_price", delimiter, errors
    )
    tax_rate = _decimal_value(
        values.get("tax_rate") or "0", row_number, "tax_rate", delimiter, errors
    )
    if tax_rate is not None and tax_rate > 100:
        errors.append(_error(row_number, "tax_rate", "KDV oranı 0 ile 100 arasında olmalıdır."))

    preparation_minutes: int | None = None
    raw_minutes = values.get("preparation_minutes", "")
    if raw_minutes:
        try:
            preparation_minutes = int(raw_minutes)
        except ValueError:
            errors.append(
                _error(row_number, "preparation_minutes", "Hazırlama süresi tam sayı olmalıdır.")
            )
        else:
            if not 0 <= preparation_minutes <= 1440:
                errors.append(
                    _error(
                        row_number,
                        "preparation_minutes",
                        "Hazırlama süresi 0 ile 1440 arasında olmalıdır.",
                    )
                )

    bool_values: dict[str, bool] = {}
    for key, default in (
        ("is_active", True),
        ("is_available", True),
        ("qr_visible", True),
        ("waiter_visible", True),
        ("track_inventory", False),
    ):
        bool_values[key] = _bool_value(values.get(key, ""), row_number, key, default, errors)

    limited_values = {
        "internal_name": 160,
        "sku": 80,
        "station": 100,
    }
    for key, maximum in limited_values.items():
        value = values.get(key, "")
        if len(value) > maximum:
            errors.append(
                _error(
                    row_number, key, f"{_display_name(key)} en fazla {maximum} karakter olabilir."
                )
            )

    if errors or selling_price is None or cost_price is None or tax_rate is None:
        return None, errors
    return (
        ParsedProductCsvRow(
            row_number=row_number,
            category=category,
            name=name,
            selling_price=selling_price,
            internal_name=values.get("internal_name") or None,
            description=values.get("description") or None,
            sku=values.get("sku") or None,
            cost_price=cost_price,
            tax_rate=tax_rate,
            preparation_minutes=preparation_minutes,
            station=values.get("station") or None,
            **bool_values,
        ),
        [],
    )


def _decimal_value(
    raw_value: str,
    row_number: int,
    field_name: str,
    delimiter: str,
    errors: list[ProductCsvImportError],
) -> Decimal | None:
    value = raw_value.strip()
    if delimiter != ",":
        value = value.replace(",", ".")
    try:
        number = Decimal(value)
    except (InvalidOperation, ValueError):
        errors.append(
            _error(
                row_number, field_name, f"{_display_name(field_name)} geçerli bir sayı olmalıdır."
            )
        )
        return None
    if not number.is_finite() or number < 0:
        errors.append(
            _error(
                row_number,
                field_name,
                f"{_display_name(field_name)} sıfır veya daha büyük olmalıdır.",
            )
        )
        return None
    exponent = number.as_tuple().exponent
    if isinstance(exponent, int) and exponent < -2:
        errors.append(
            _error(
                row_number,
                field_name,
                f"{_display_name(field_name)} en fazla iki ondalık basamak içerebilir.",
            )
        )
        return None
    if number >= Decimal("1000000000000"):
        errors.append(
            _error(
                row_number, field_name, f"{_display_name(field_name)} izin verilen sınırı aşıyor."
            )
        )
        return None
    return number.quantize(Decimal("0.01"))


def _bool_value(
    raw_value: str,
    row_number: int,
    field_name: str,
    default: bool,
    errors: list[ProductCsvImportError],
) -> bool:
    value = _normalize(raw_value)
    if not value:
        return default
    if value in _TRUE_VALUES:
        return True
    if value in _FALSE_VALUES:
        return False
    errors.append(
        _error(
            row_number, field_name, f"{_display_name(field_name)} için evet veya hayır kullanın."
        )
    )
    return default


def _normalize(value: str) -> str:
    # Unicode decomposition does not transliterate Turkish dotless-i. Map it
    # explicitly so values such as "hayır" match the ASCII aliases/booleans.
    normalized = unicodedata.normalize("NFKD", value.strip().casefold().replace("ı", "i"))
    ascii_value = "".join(
        character for character in normalized if not unicodedata.combining(character)
    )
    return re.sub(r"[^a-z0-9]+", "_", ascii_value).strip("_")


def normalize_lookup_name(value: str) -> str:
    return _normalize(value)


def _display_name(field_name: str) -> str:
    return {
        "category": "Kategori",
        "name": "Ürün adı",
        "selling_price": "Satış fiyatı",
        "cost_price": "Maliyet fiyatı",
        "tax_rate": "KDV oranı",
        "preparation_minutes": "Hazırlama süresi",
        "internal_name": "Dahili ad",
        "station": "İstasyon",
        "sku": "SKU",
        "is_active": "Aktif",
        "is_available": "Satışa açık",
        "qr_visible": "QR menü görünürlüğü",
        "waiter_visible": "Garson menüsü görünürlüğü",
        "track_inventory": "Stok takibi",
    }.get(field_name, field_name)


def _error(row_number: int, field_name: str | None, message: str) -> ProductCsvImportError:
    return ProductCsvImportError(row_number=row_number, field=field_name, message=message)
