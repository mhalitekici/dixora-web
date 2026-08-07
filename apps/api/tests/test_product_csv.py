from __future__ import annotations

import io
from decimal import Decimal

from openpyxl import load_workbook

from app.services.product_csv import (
    build_product_csv_template,
    build_product_xlsx_template,
    parse_product_csv,
    parse_product_xlsx,
)
from tests.conftest import ApiContext, auth_headers, login


async def test_csv_template_requires_auth_and_is_utf8_bom(api: ApiContext) -> None:
    unauthorized = await api.client.get("/api/v1/catalog/products/csv-template")
    assert unauthorized.status_code == 401

    owner = await login(api)
    response = await api.client.get(
        "/api/v1/catalog/products/csv-template",
        headers=auth_headers(owner),
    )
    assert response.status_code == 200, response.text
    assert response.content.startswith(b"\xef\xbb\xbf")
    assert "dixora-urun-sablonu.csv" in response.headers["content-disposition"]
    template = response.content.decode("utf-8-sig")
    assert template.startswith("sep=;\r\nkategori;urun_adi;satis_fiyati")


async def test_xlsx_template_is_native_excel_table_and_requires_auth(api: ApiContext) -> None:
    unauthorized = await api.client.get("/api/v1/catalog/products/import-template")
    assert unauthorized.status_code == 401

    owner = await login(api)
    response = await api.client.get(
        "/api/v1/catalog/products/import-template",
        headers=auth_headers(owner),
    )
    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "dixora-urun-sablonu.xlsx" in response.headers["content-disposition"]

    workbook = load_workbook(io.BytesIO(response.content))
    assert workbook.sheetnames == ["Ürünler", "Nasıl Kullanılır"]
    products = workbook["Ürünler"]
    assert products.freeze_panes == "A2"
    assert products["A1"].value == "Kategori"
    assert products["B1"].value == "Ürün Adı"
    assert products["B2"].value == "Izgara Köfte"
    assert products.tables["DixoraUrunTablosu"].ref == "A1:O2001"
    assert len(products.data_validations.dataValidation) == 3
    assert workbook["Nasıl Kullanılır"]["A1"].value == "Dixora ürün aktarım şablonu"
    workbook.close()


def test_excel_template_round_trips_with_turkish_decimal_values() -> None:
    parsed = parse_product_csv(build_product_csv_template())

    assert parsed.errors == []
    assert parsed.total_rows == 1
    assert len(parsed.rows) == 1
    assert parsed.rows[0].row_number == 3
    assert parsed.rows[0].name == "Izgara Köfte"
    assert parsed.rows[0].selling_price == Decimal("325.00")
    assert parsed.rows[0].cost_price == Decimal("120.00")


def test_xlsx_template_round_trips_with_unicode_and_native_numbers() -> None:
    parsed = parse_product_xlsx(build_product_xlsx_template())

    assert parsed.errors == []
    assert parsed.total_rows == 1
    assert parsed.rows[0].row_number == 2
    assert parsed.rows[0].category == "Ana Yemekler"
    assert parsed.rows[0].name == "Izgara Köfte"
    assert parsed.rows[0].description == "Pilav ve mevsim salatası ile"
    assert parsed.rows[0].selling_price == Decimal("325.00")
    assert parsed.rows[0].track_inventory is False


def test_xlsx_parser_rejects_formulas_and_invalid_workbooks_safely() -> None:
    workbook = load_workbook(io.BytesIO(build_product_xlsx_template()))
    products = workbook["Ürünler"]
    products["C2"] = "=100+50"
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()

    parsed = parse_product_xlsx(output.getvalue())
    assert parsed.total_rows == 1
    assert parsed.rows == []
    assert any("Formül kullanılamaz" in error.message for error in parsed.errors)

    invalid = parse_product_xlsx(b"not-an-xlsx-file")
    assert invalid.rows == []
    assert invalid.errors[0].message == (
        "Excel dosyası okunamadı. Geçerli bir .xlsx dosyası yükleyin."
    )


def test_csv_parser_keeps_accepting_standard_comma_delimited_files() -> None:
    parsed = parse_product_csv(
        (
            "category,name,selling_price,description,is_active\r\n"
            'Burgers,"Burger, Patates",199.90,"Acılı, büyük porsiyon",yes\r\n'
        ).encode()
    )

    assert parsed.errors == []
    assert parsed.total_rows == 1
    assert parsed.rows[0].row_number == 2
    assert parsed.rows[0].name == "Burger, Patates"
    assert parsed.rows[0].selling_price == Decimal("199.90")
    assert parsed.rows[0].description == "Acılı, büyük porsiyon"


def test_csv_parser_honors_excel_comma_separator_directive() -> None:
    parsed = parse_product_csv(
        b"sep=,\r\ncategory,name,selling_price,is_active\r\nBurgers,Classic Burger,179.50,yes\r\n"
    )

    assert parsed.errors == []
    assert parsed.rows[0].row_number == 3
    assert parsed.rows[0].selling_price == Decimal("179.50")


def test_csv_parser_defaults_blank_optional_numeric_cells_to_zero() -> None:
    parsed = parse_product_csv(
        (
            "kategori;urun_adi;satis_fiyati;maliyet_fiyati;kdv_orani\r\n"
            "İçecekler;Filtre Kahve;95,00;;\r\n"
        ).encode()
    )

    assert parsed.errors == []
    assert parsed.rows[0].cost_price == Decimal("0.00")
    assert parsed.rows[0].tax_rate == Decimal("0.00")


async def test_csv_preview_reports_errors_and_imports_only_valid_rows(api: ApiContext) -> None:
    owner = await login(api)
    headers = auth_headers(owner)
    csv_data = (
        "kategori;urun_adi;satis_fiyati;sku;aktif\r\n"
        "Burgers;CSV Test Burger;199,90;CSV-TEST-001;evet\r\n"
        "Olmayan Kategori;Hatalı Ürün;40,00;CSV-TEST-002;evet\r\n"
    ).encode()

    preview = await api.client.post(
        "/api/v1/catalog/products/csv-import",
        headers=headers,
        params={"dry_run": "true"},
        files={"file": ("urunler.csv", csv_data, "text/csv")},
    )
    assert preview.status_code == 200, preview.text
    assert preview.json()["status"] == "PARTIAL"
    assert preview.json()["total_rows"] == 2
    assert preview.json()["valid_rows"] == 1
    assert preview.json()["failed_rows"] == 1
    assert preview.json()["imported_rows"] == 0
    assert preview.json()["errors"][0]["row_number"] == 3

    imported = await api.client.post(
        "/api/v1/catalog/products/csv-import",
        headers=headers,
        params={"dry_run": "false"},
        files={"file": ("urunler.csv", csv_data, "text/csv")},
    )
    assert imported.status_code == 200, imported.text
    assert imported.json()["status"] == "PARTIAL"
    assert imported.json()["imported_rows"] == 1

    products = await api.client.get(
        "/api/v1/catalog/products",
        headers=headers,
        params={"search": "CSV Test Burger"},
    )
    assert products.status_code == 200, products.text
    assert [item["name"] for item in products.json()["items"]] == ["CSV Test Burger"]


async def test_xlsx_preview_and_import_support_turkish_product_data(api: ApiContext) -> None:
    owner = await login(api)
    headers = auth_headers(owner)
    workbook = load_workbook(io.BytesIO(build_product_xlsx_template()))
    products_sheet = workbook["Ürünler"]
    products_sheet["A2"] = "Burgers"
    products_sheet["B2"] = "Köz Biberli Köfte"
    products_sheet["C2"] = 249.9
    products_sheet["E2"] = "İsli peynir, köz biber ve özel sos"
    products_sheet["F2"] = "XLSX-TR-001"
    products_sheet["J2"] = None
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()

    preview = await api.client.post(
        "/api/v1/catalog/products/csv-import",
        headers=headers,
        params={"dry_run": "true"},
        files={
            "file": (
                "urunler.xlsx",
                output.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert preview.status_code == 200, preview.text
    assert preview.json()["status"] == "READY"
    assert preview.json()["valid_rows"] == 1
    assert preview.json()["rows"][0]["name"] == "Köz Biberli Köfte"

    imported = await api.client.post(
        "/api/v1/catalog/products/csv-import",
        headers=headers,
        params={"dry_run": "false"},
        files={
            "file": (
                "urunler.xlsx",
                output.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert imported.status_code == 200, imported.text
    assert imported.json()["status"] == "SUCCESS"
    assert imported.json()["imported_rows"] == 1
