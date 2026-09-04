"""Exporting the catalogue, and getting it back in unchanged.

The point of the export is to move a menu into another business, so the test
that really matters is the round trip: download, read it back, compare.
"""

from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from uuid import uuid4

from openpyxl import load_workbook
from sqlalchemy import select

from app.models import Product, Tenant
from app.services.product_csv import (
    XLSX_TEMPLATE_HEADERS,
    build_product_export,
    parse_product_xlsx,
)
from tests.conftest import ApiContext, auth_headers, login


async def test_the_export_downloads_as_a_workbook(api: ApiContext) -> None:
    headers = auth_headers(await login(api))
    response = await api.client.get("/api/v1/catalog/products/export", headers=headers)
    assert response.status_code == 200, response.text
    assert "spreadsheetml" in response.headers["content-type"]
    assert "dixora-urunler-" in response.headers["content-disposition"]
    # A real xlsx is a zip archive; "PK" is its magic number.
    assert response.content[:2] == b"PK"


async def test_the_export_carries_every_product(api: ApiContext) -> None:
    headers = auth_headers(await login(api))
    response = await api.client.get("/api/v1/catalog/products/export", headers=headers)
    parsed = parse_product_xlsx(response.content)

    async with api.database.session_factory() as db:
        products = (await db.execute(select(Product))).scalars().all()

    assert len(parsed.rows) == len(products)
    assert {row.name for row in parsed.rows} == {p.name for p in products}


async def test_the_export_uses_the_layout_the_importer_expects(
    api: ApiContext,
) -> None:
    """Export and template must not drift apart, or the round trip breaks."""
    async with api.database.session_factory() as db:
        tenant = (await db.execute(select(Tenant))).scalars().first()
        assert tenant is not None
        content = await build_product_export(db, tenant_id=tenant.id)

    sheet = load_workbook(BytesIO(content))["Ürünler"]
    assert tuple(cell.value for cell in sheet[1]) == XLSX_TEMPLATE_HEADERS


async def test_an_exported_menu_reads_back_with_prices_intact(
    api: ApiContext,
) -> None:
    """The whole reason the export exists: load this file into a new business."""
    headers = auth_headers(await login(api))
    exported = await api.client.get(
        "/api/v1/catalog/products/export", headers=headers
    )
    assert exported.status_code == 200

    async with api.database.session_factory() as db:
        source = (await db.execute(select(Tenant))).scalars().first()
        assert source is not None
        original = {
            p.name: (p.selling_price, p.sku)
            for p in (
                await db.execute(select(Product).where(Product.tenant_id == source.id))
            )
            .scalars()
            .all()
        }

    parsed = parse_product_xlsx(exported.content)
    assert parsed.rows, "export produced no rows"
    assert not parsed.errors, parsed.errors

    for row in parsed.rows:
        assert row.name in original
        price, sku = original[row.name]
        assert Decimal(str(row.selling_price)) == price
        assert (row.sku or None) == sku


async def test_the_export_never_reaches_another_business(api: ApiContext) -> None:
    async with api.database.session_factory() as db:
        other = Tenant(
            name="Rakip",
            slug=f"rakip-{uuid4().hex[:8]}",
            business_type="CAFE",
            state="ACTIVE",
            is_active=True,
        )
        db.add(other)
        await db.commit()
        content = await build_product_export(db, tenant_id=other.id)

    parsed = parse_product_xlsx(content)
    # A business with no catalogue exports an empty sheet, not everyone else's.
    assert parsed.rows == []


async def test_the_export_requires_authentication(api: ApiContext) -> None:
    response = await api.client.get("/api/v1/catalog/products/export")
    assert response.status_code == 401


async def _upload(
    api: ApiContext,
    headers: dict[str, str],
    content: bytes,
    *,
    dry_run: bool,
    update_existing: bool,
):
    return await api.client.post(
        "/api/v1/catalog/products/csv-import"
        f"?dry_run={str(dry_run).lower()}&update_existing={str(update_existing).lower()}",
        headers=headers,
        files={
            "file": (
                "urunler.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )


async def test_reimporting_without_update_mode_still_refuses(
    api: ApiContext,
) -> None:
    """The default must keep protecting a live menu from being overwritten."""
    headers = auth_headers(await login(api))
    async with api.database.session_factory() as db:
        # Seeded products carry no stock codes, and the refusal is keyed on
        # them, so one has to exist for this rule to be exercised at all.
        product = (await db.execute(select(Product))).scalars().first()
        assert product is not None
        product.sku = "TEST-SKU-1"
        await db.commit()

    exported = await api.client.get(
        "/api/v1/catalog/products/export", headers=headers
    )
    response = await _upload(
        api, headers, exported.content, dry_run=True, update_existing=False
    )
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["failed_rows"] > 0
    assert any("zaten kullanılıyor" in e["message"] for e in body["errors"])


async def test_update_mode_rewrites_the_matching_products(
    api: ApiContext,
) -> None:
    """Edit prices in a spreadsheet, upload it back, see the new prices."""
    headers = auth_headers(await login(api))
    exported = await api.client.get(
        "/api/v1/catalog/products/export", headers=headers
    )

    workbook = load_workbook(BytesIO(exported.content))
    sheet = workbook["Ürünler"]
    price_column = XLSX_TEMPLATE_HEADERS.index("Satış Fiyatı") + 1
    name_column = XLSX_TEMPLATE_HEADERS.index("Ürün Adı") + 1
    edited_name = sheet.cell(row=2, column=name_column).value
    sheet.cell(row=2, column=price_column).value = 999.5
    buffer = BytesIO()
    workbook.save(buffer)

    async with api.database.session_factory() as db:
        before = (
            await db.execute(select(Product).where(Product.name == edited_name))
        ).scalar_one()
        assert before.selling_price != Decimal("999.50")
        product_id = before.id
        count_before = len((await db.execute(select(Product))).scalars().all())

    response = await _upload(
        api, headers, buffer.getvalue(), dry_run=False, update_existing=True
    )
    assert response.status_code == 200, response.text
    assert response.json()["failed_rows"] == 0

    async with api.database.session_factory() as db:
        after = await db.get(Product, product_id)
        assert after is not None
        assert after.selling_price == Decimal("999.50")
        # Updated in place — the catalogue did not double.
        count_after = len((await db.execute(select(Product))).scalars().all())
        assert count_after == count_before
